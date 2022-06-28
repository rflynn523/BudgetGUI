# File that runs the checks after New Month is pressed and shows the results to the user
import tkinter as tk
import openpyxl as xl
import info, newMonth

# Adds all the labels to the GUI and then calls all the check functions
def makeAllChecks():
    # Current Checks that are going to run
    checkList = ["Monthly Amounts Table Reset", "Current Month Check on Monthly", "Amounts Entered into Yearly",
                 "Totals entered into Yearly", "Category Totals reset on Monthly",
                 "Entry Table Emptied", "Entries entered into Data Set"]

    # Add those labels to the new window
    for x in range(len(checkList)):
        tk.Label(checkWindow, text=checkList[x] + ":", font="Calibri 12").grid(row=x, column=0,columnspan=1, sticky=tk.W, padx=5, pady=5)

    # Relaod the workbooks
    wbData = xl.load_workbook(info.excelFile, data_only=True)
    wbEq = xl.load_workbook(info.excelFile, data_only=False)

    # Get the needed sheets
    monthSheetData = wbData['Monthly']
    yearSheetData = wbData['Yearly']
    dataSetSheetData = wbData["Data Set"]

    monthSheetEq = wbEq["Monthly"]
    yearSheetEq = wbEq["Yearly"]
    dataSetSheetEq = wbEq["Data Set"]

    # Make all the checks
    # Check that the amounts table on monthly is 0 for non fixed categories
    monthlyTableReset(monthSheetData)

    # Check that the correct month is displayed on Monthly
    currentMonthCheck(monthSheetData)

    # Check that the amounts were entered into the correct month on Yearly
    # amountsIntoYearly(monthSheetData, yearSheetData)

    # Check that the 'old' months Total Spent, and NET values are
    # inserted to the table on Yearly.
    totalsIntoYearly(yearSheetData)

    # Check that the 'new' month cells are the actual equations
    # monthlyTotalsReset(monthSheetEq)

    # Make sure the Entry table is empty
    emptyEntryTable(monthSheetData)

    # Make sure the data was inserted into the Data Set sheet with the MONTH() formula
    # dataSetEntered(dataSetSheetData, yearSheetData)

    # Finally show the new window to the user
    showGUI()

# Check that the amounts table on monthly is 0 for non fixed categories
def monthlyTableReset(monthSheetData):
    col = 2

    # List of row numbers that contain equation based totals
    equationCellRows = [8, 10, 11, 12, 14, 15]

    # Check each cell to see if the data is None/0 mean it was reset correctly
    for cellRow in equationCellRows:
        if(monthSheetData[cellRow][col].value != None):
            addResult(False, 0)
            return

    # If you make it through the for loop then they are all good
    addResult(True, 0)

# Check that the correct month is displayed on Monthly
def currentMonthCheck(monthSheetData):
    monthCheck = (info.month == monthSheetData.cell(row = 23, column = 1).value)
    addResult(monthCheck, 1)

# Check that the amounts were entered into the correct month on Yearly
def amountsIntoYearly(monthSheetData, yearSheetData):
    numCategories = len(info.categoryList)

    # Get the next open row which is now below the month that was just written to yearly
    nextOpenRow = info.getRowNum(yearSheetData, 23, 2)

    # Start looking at the beggining of the categories
    currentRow = nextOpenRow - numCategories

    matchesMonthlyAmounts = []

    # Loop through the data on yearly
    # None corresponds to no data in the cell
    while currentRow <= nextOpenRow:
        matchesMonthlyAmounts.append(yearSheetData[currentRow][2].value == None)
        currentRow += 1

        # If a false is found return right away
        if (False in matchesMonthlyAmounts):
            addResult(False, 2)
            return

    addResult(True, 2)

# Check that the 'old' months Total Spent, and NET values are
# inserted to the table on Yearly.
def totalsIntoYearly(yearSheetData):
    # With the zero indexed list, info.month is the month number of the previous month
    monthRow = info.getRowNum(yearSheetData, 4, 3) - 1
    print(monthRow)

    totalSpent = (yearSheetData.cell(row=monthRow-1, column=3).value != None)
    net = (yearSheetData.cell(row=monthRow-1, column=4).value != None)

    addResult((totalSpent and net), 3)

# Check that the 'new' month cells are the actual equations
def monthlyTotalsReset(monthSheetEq):
    # Find the last row in the monthly category table
    row = 5
    col = 2
    lastRow = info.getRowNum(monthSheetEq, row, 1)

    # List of row numbers that contain equation based totals
    equationCellRows_Expenses = [3, 6, 7, 8, 9, 10, 11, 12, 14, 15]
    equationCellRows_Apartment = [8, 9]

    print(info.isApartment)

    # Check to see which file si loaded bc each file has diffrent rows to check
    if(info.isApartment == True):
        equationCellRows = equationCellRows_Apartment
    else:
        equationCellRows = equationCellRows_Expenses

    print(equationCellRows)
    
    # Check each cell to see if the data is None/0 mean it was reset correctly
    for cellRow in equationCellRows:
        print(monthSheetEq[cellRow][col].value)
        if(monthSheetEq[cellRow][col].value[0] != "="):
            addResult(False, 4)
            return

    addResult(True, 4)

# Make sure the Entry table is empty
def emptyEntryTable(monthSheetData):
    addResult((info.getRowNum(monthSheetData, 25, 1) == 25), 5)

 # Make sure the data was inserted into the Data Set sheet
def dataSetEntered(dataSetSheetData, yearSheetData):
    # Get the first row number that has the index of the month
    if (info.month == "January"):
        firstRow = info.getRowNum(dataSetSheetData, 3, 3, month=12) + 1
    else:
        firstRow = info.getRowNum(dataSetSheetData, 3, 4, month=(info.months.index(info.month)) + 1)

    # Get the last row number
    lastRow = info.getRowNum(dataSetSheetData, firstRow, 4) - 1

    print(lastRow)
    print(firstRow)
    print(len(info.categoryList))

    # Simply check if the number of rows in Data Set with the current month
    # match the number of entries that was saved while clearing the entries table
    addResult(((lastRow - firstRow) == len(info.categoryList)), 6)

# Helper funciton that adds the results to the pop-up window
def addResult(result, num):
    if result:
        text = "Success"
    else:
        text = "Fail"

    # Add the result label to the windowS
    tk.Label(checkWindow, text=text, font="Calibri 12 bold").grid(row=num, column=1, columnspan=1, sticky=tk.W,
                                                                      padx=5, pady=5)

# After all the checks have been complete, show the window to the user
def showGUI():
    # Add Close button that just closes the window
    tk.Button(checkWindow, text="Close", font="Calibri 12 bold", relief='groove', bg="mediumseagreen",
              activebackground="darkolivegreen", command=checkWindow.destroy).grid(row=10, column=0, columnspan=2, sticky=tk.S+tk.W+tk.E, padx=5, pady=5)

    checkWindow.update()
    checkWindow.deiconify()
    checkWindow.geometry('+600+100')
    checkWindow.title("New Month Check")

# Window
checkWindow = tk.Toplevel(info.window)
checkWindow.withdraw()
