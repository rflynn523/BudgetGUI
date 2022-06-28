# Contains information and some helper functions that almost every file needs.

import openpyxl as xl
import tkinter as tk

from tkinter import filedialog as fd

import createGUI
from openpyxl.styles import Font, Fill, PatternFill, Border, Side

# Helper Function
# If key is not passed then the function returns the open row number
# Otherwise it returns the row that contains the key
def getRowNum(sheet, startRow, startCol, key=None, month=None):
    # Simply return the next open row number with the given params
    if(key == None and month==None):
        while(sheet[startRow][startCol].value != None):
            startRow += 1
            if(startRow == 2000):
                break

    # Return the row where the cell is equal to the key
    elif(month == None):
        # Makes sure program doesnt crash if string is handled wrong
        if(type(key) == str):
            while (str(sheet[startRow][startCol].value) != key):
                startRow += 1
                if (startRow == 2000):
                    break
        else:
            while (sheet[startRow][startCol].value != key):
                startRow += 1
                if (startRow == 2000):
                    break

    # Used for the Data Set Check
    # Returns the row when the date is equal to the month (passed in as a number)
    else:
        try:
            while ((sheet[startRow+1][startCol].value).month != month):
                startRow += 1
                if (startRow == 100):
                    break
        except:
            createGUI.displayMessage("StartRow error? info.py line 41")

    return startRow

# Function to save a Back up file before the script makes changes to the main file
def createBackUpFile(newFileName):
    wbEq.save(newFileName)

# Function that opens the file dialog box so you can select the Excel file
def select_file():
    fileTypes = (('excel files', '*.xlsx'), ('All files', '*.*'))

    fileName = fd.askopenfilename(
        title = 'Select Excel file',
        initialdir = 'C:/Users/rflyn/Desktop/BudgetGUI/',
        filetypes = fileTypes
    )

    print(fileName)

    fileNameStripped = fileName.split("/")
    return fileNameStripped[-1]


# Window variable
window = tk.Tk()

excelFile = select_file()

# Formatting and other info
accountingFormat = r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)'
dateFormat = 'dd-mmm'
noFill = PatternFill(fill_type=None)

side = Side(border_style=None)
noBorder = Border(
    left=side,
    right=side,
    top=side,
    bottom=side,
)

thin = Side(border_style='thin')
allBorders = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin,
)
 
lightGreenFill = PatternFill(fill_type='solid',
                 start_color='C6E0B4',
                 end_color='C6E0B4')
                 

greenFill = PatternFill(fill_type='solid',
                 start_color='00B050',
                 end_color='00B050')

yearlyMonth = Font(name="Arial Black", bold = True)

boldFont = Font(bold = True)

months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

# # Cells correspond to the above months and contain the first cell under 'Amount' on the Yearly sheet
# # Updated the cells on 5/2/21
# 
# These cells are not needed as of 6/27/2022, changed the format of the yearly sheet to make it easier to make changes to new or old categories
# Keeping for now just in case I need to go back, but eventually can be REMOVED
# 
# ApartmentCells = [[5,3], [5,6], [5,9], [17,3], [17,6], [17,9], [29,3], [29,6], [29,9], [41,3], [41,6], [41,9]]
# ExpenseCells = [[5,3], [5,6], [5,9], [21,3], [21,6], [21,9], [37,3], [37,6], [37,9], [53,3], [53,6], [53,9]]

# Assume that the file is the Expenses
# REMOVE but also need to remove from newMonthCheck
isApartment = False


# Load the workbooks
wbData = xl.load_workbook(excelFile, data_only=True)
wbEq = xl.load_workbook(excelFile, data_only=False)

# Get the needed sheets
monthSheetData = wbData['Monthly']
yearSheetData = wbData['Yearly']
dataSetSheetData = wbData["Data Set"]

monthSheetEq = wbEq["Monthly"]
yearSheetEq = wbEq["Yearly"]
dataSetSheetEq = wbEq["Data Set"]

# Initialize these variables to be used later
month = monthSheetData[23][0].value

categoryList = []
numEntries = 0

# Loops through the category names starting with the rent cell
# Load in the different categories
row = 3
col = 2
lastRow = getRowNum(monthSheetEq, row, 1)

for x in range(lastRow - row):
    categoryList.append(monthSheetData[row][1].value)
    row += 1
