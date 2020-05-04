# Handles all of the nessesary steps and functions need to reset the spreadsheet for a new month
import info
import openpyxl as xl

def new_month():
    # Copy the VALUES from the monthly Category Table to the corresponding table in Yearly
    copySummaryTable(info.monthSheetData, info.yearSheetEq)

    # Also copy the Total Spent, Total Besides P/R, and Net values from Monthly to Yearly
    copyTotalValues(info.monthSheetData, info.yearSheetData, info.yearSheetEq)

    # Update the Grocery and Gas tables by moving the equation to the next month and only
    # saving the values of the previous month.
    updateGrocGasTables(info.monthSheetData, info.monthSheetEq)

    # Get the data from the Entries table from Monthly, write that data to the 'Data Set'
    # sheet and then clear the Entries table
    updateEntryTable(info.monthSheetData, info.monthSheetEq, info.dataSetSheetEq)

    # Save only the EQUATIONS workbook file
    info.wbEq.save(info.excelFile)

    # Call the updateValues function to update the GUI once its complete

def getMonthStartCell():
    for i in range(len(info.months)):
        check = info.months[i]
        if(info.month == check):
            # Update month

            # Loop around to the start of the list
            if (i == 11):
                i = 0

            updateMonth(info.months[i+1])
            return info.yearly_month_cells.get(check)

def updateMonth(newMonth):
    with open(r"BudgetGuiConfig.txt", 'r') as oldFile:
        data = oldFile.readlines()

    data[0] = newMonth + '\n'
    data[1] = info.excelFile

    with open(r'BudgetGuiConfig.txt', 'w') as newFile:
        newFile.writelines(data)

    info.month = newMonth

    # Write the new month to the monthly cell

def copySummaryTable(monthSheetData, yearSheetEq):
    # Copy only the values from the summary table to the "Yearly" Sheet

    # Starting at 'Rent' cell in amounts column of the category table
    categoryList = []
    row = 3
    monthCell = monthSheetData[row][2]

    # Loop down the amounts column, saving all of the numbers
    for x in range(8):
        # Clean values are not needed because excel needs to see the values as numbers not strings
        categoryList.append(monthCell.value)
        row += 1
        monthCell = monthSheetData[row][2]

    # Write to the yearly sheet of the EQUATIONS file
    # For the current month according to the config file
    monthCellList = getMonthStartCell()
    row, col = monthCellList[0], monthCellList[1]

    for y in range(8):
        yearSheetEq.cell(row = row, column = col).value = categoryList[y];
        row += 1

# # # Copy Total Spent, Total(Besides R/P), and NET cells to yearly\
def copyTotalValues(monthSheetData, yearSheetData, yearSheetEq):
    # Getting data from DATA wb
    totalSpent = monthSheetData[12][2].value
    totBesidesRP = monthSheetData[14][2].value
    net =  monthSheetData[16][2].value

    nextOpen = info.getOpenRow(yearSheetData, 22, 12)

    # Writing data to EQUATION file
    yearSheetEq.cell(row=nextOpen, column=13).value = totalSpent
    yearSheetEq.cell(row=nextOpen, column=14).value = totBesidesRP
    yearSheetEq.cell(row=nextOpen, column=15).value = net

def updateGrocGasTables(monthSheetData, monthSheetEq):
    # # Copy the values in the groceries and gas tables into to the same cells so those months get saved.
    current = info.getOpenRow(monthSheetData, 31, 10) - 1

    grocTotalData = monthSheetData[current][9].value
    grocAvgData = monthSheetData[current][10].value

    grocTotalEq = monthSheetEq[current][9].value
    grocAvgEq = monthSheetEq[current][10].value

    # Write the data into its cell
    monthSheetEq.cell(row=current, column=10).value = grocTotalData
    monthSheetEq.cell(row=current, column=11).value = grocAvgData

    # Move the equation to the cell below
    monthSheetEq.cell(row=current + 1, column=10).value = grocTotalEq
    monthSheetEq.cell(row=current + 1, column=11).value = grocAvgEq

    # Do the same thing for the gas table
    gasTotalData = monthSheetData[current][13].value
    gasAvgData = monthSheetData[current][14].value

    gasTotalEq = monthSheetEq[current][13].value
    gasAvgEq = monthSheetEq[current][14].value

    # Write the data into its cell
    monthSheetEq.cell(row=current, column=14).value = gasTotalData
    monthSheetEq.cell(row=current, column=15).value = gasAvgData

    # Move the equation to the cell below
    monthSheetEq.cell(row=current + 1, column=14).value = gasTotalEq
    monthSheetEq.cell(row=current + 1, column=15).value = gasAvgEq

def updateEntryTable(monthSheetData, monthSheetEq, dataSetSheet):
# Get the last row number of the data
    lastRow = info.getOpenRow(monthSheetData, 24, 1) - 1
    print(lastRow)
    dataList = []

    # Get each data row into a list and then add that list to dataList
    for i in range(25, lastRow + 1):
        dataRow = []
        for j in range(1, 6):
            dataRow.append(monthSheetData.cell(row = i, column = j).value)

        dataList.append(dataRow)

    # Copy the dataList into the Data Set sheet and insert the MONTH() formula
    nextOpen = info.getOpenRow(dataSetSheet, 3, 4)

    # Insert the data row by row
    for d in range(nextOpen, nextOpen + len(dataList)):
        temp = dataList[d - nextOpen]
        dataSetSheet.cell(row=d, column=4).value = "=MONTH(E" + str(d) + ")"
        for c in range(5, 10):
            dataSetSheet.cell(row = d, column = c).value = temp[c - 5]

            # Date format
            if (c == 5):
                dataSetSheet.cell(row=d, column=c).number_format = info.dateFormat

            # Money format
            if (c == 8):
                dataSetSheet.cell(row=d, column=c).number_format = info.accountingFormat

    # Clear the entries table from the "Monthly" sheet
    for i in range(25, lastRow + 1):
        for j in range(1, 6):
            info.monthSheetEq.cell(row = i, column = j).value = None